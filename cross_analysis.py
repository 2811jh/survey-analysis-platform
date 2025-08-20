import pandas as pd
import re
import os
import warnings
import numpy as np
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.chart import BarChart, Reference
from scipy.stats import chi2_contingency, fisher_exact

def extract_subcol_number(subcol, prefix):
    suffix = subcol.split(prefix)[1].strip()
    match = re.search(r'^(\d+)', suffix)
    return int(match.group(1)) if match else 0

def perform_significance_test(observed):
    """执行统计检验并返回p值"""
    try:
        # 尝试卡方检验
        chi2, p, dof, expected = chi2_contingency(observed)
        # 检查期望频数假设（超过20%的单元格期望频数小于5时使用费舍尔精确检验）
        if (expected < 5).sum() / expected.size > 0.2:
            _, p = fisher_exact(observed)
    except:
        # 当出现计算错误时使用费舍尔精确检验
        try:
            _, p = fisher_exact(observed)
        except:
            p = np.nan
    return p

def process_crosstab(
    input_file, 
    output_file, 
    row_questions, 
    col_questions,
    # 新增显著性检验参数
    sig_levels=[0.05, 0.01, 0.001],
    sig_symbols=['*', '**', '***'],
    # 样式配置
    header_height=55,
    header_fill_color="4F81BD",
    header_font_color="FFFFFF",
    header_font_name="微软雅黑",
    header_font_size=12,
    # 数据条配置
    freq_data_bar_color="638EC6",
    percent_data_bar_color="C00000",
    data_bar_min_length=15,
    data_bar_max_length=100,
    # 格式配置
    percent_format="0.00%",
    data_column_width=20,  # C列及之后的固定宽度（None表示自动调整）
    max_column_width=40 #【可调整最大列宽】
):
    
    # === 数据准备 ===
    try:
        df = pd.read_excel(input_file)
        df.columns = [str(col).strip() for col in df.columns]  # 统一清理列名
        
        # ===== 新增：变量重新编码功能 =====
        def recode_variable(df, source_col, new_col, rules):
            """
            重新编码变量
            参数：
            - df: 原始DataFrame
            - source_col: 原始列名
            - new_col: 新列名
            - rules: 编码规则字典
            """
            if source_col not in df.columns:
                warnings.warn(f"原始列 {source_col} 不存在，跳过重新编码")
                return df
            
            if rules['method'] == 'map':
                df[new_col] = df[source_col].map(rules['config']['mapping'])
            elif rules['method'] == 'cut':
                df[new_col] = pd.cut(
                    df[source_col], 
                    bins=rules['config']['bins'],
                    labels=rules['config']['labels'],
                    ordered=False
                )
            else:
                raise ValueError("不支持的编码方式")
            
            return df
        
        # ===== 编码规则配置 =====
        recode_rules = []
        
        # 应用所有编码规则
        for rule in recode_rules:
            df = recode_variable(df, **rule)
    
    except Exception as e:
        raise Exception(f"输入文件读取失败: {str(e)}")

    # === 识别用户配置的多选题根 ===
    user_multi_roots = set()
    for q in row_questions + col_questions:
        q_clean = str(q).strip()
        if re.fullmatch(r'^Q\d+\.$', q_clean):  # 严格匹配 Q数字. 格式
            user_multi_roots.add(q_clean)
    
    # === 构建多选题字典 ===
    multi_choice_dict = defaultdict(list)
    for col in df.columns:
        col_clean = str(col).strip()
        for root in user_multi_roots:
            if col_clean.startswith(root):
                multi_choice_dict[root].append(col)
    
    # 排序并过滤有效多选题
    multi_choice_dict = {
        k: sorted(v, key=lambda x: extract_subcol_number(x, k))
        for k, v in multi_choice_dict.items() if len(v) > 1
    }
    
    # === 验证问题有效性 ===
    def validate_questions(questions):
        valid = []
        invalid = []
        for q in questions:
            q_clean = str(q).strip()

            if q_clean in df.columns:
                valid.append(('single', q_clean))

            elif q_clean in multi_choice_dict:
                valid.append(('multi', q_clean))

            else:
                invalid.append(q)
        return valid, invalid
    
    valid_rows, invalid_rows = validate_questions(row_questions)
    valid_cols, invalid_cols = validate_questions(col_questions)
    
    if invalid_rows or invalid_cols:
        warnings.warn(f"无效问题将被跳过：行问题={invalid_rows}，列问题={invalid_cols}")

    # === 问题有效性验证 ===
    invalid_questions = []
    valid_rows = []

    # 处理行问题
    for q in row_questions:
        found = False
        if re.match(r'^Q\d+\.', str(q)):
            root = re.match(r'^(Q\d+\.)', str(q)).group(1)
            if root in multi_choice_dict:
                valid_rows.append(('multi', root))
                found = True
        if not found and q in df.columns:
            valid_rows.append(('single', q))
            found = True
        if not found:
            invalid_questions.append(q)

    # === 列条件生成 ===
    col_conditions = []
    col_totals = {}
    seen_cols = defaultdict(int)  # 记录列问题出现次数
    
    for q in col_questions:  # 保留原始顺序，不跳过重复项
        q_clean = str(q).strip()
        seen_cols[q_clean] += 1
        instance_id = seen_cols[q_clean]
    
        # === 处理多选题 ===
        if re.match(r'^Q\d+\.', q_clean):
            root = re.match(r'^(Q\d+\.)', q_clean).group(1)
            if root in multi_choice_dict:
                subcols = sorted(multi_choice_dict[root], key=lambda x: extract_subcol_number(x, root))
                example_subcol = subcols[0]
                rest_part = example_subcol.split(root)[1].strip()
                if ':' in rest_part:
                    question_text, _ = rest_part.split(':', 1)
                    question_text = question_text.strip()
                else:
                    question_text = rest_part
                full_question = f"{root}{question_text} #{instance_id}"  # 唯一标识
    
                for subcol in subcols:
                    rest_subcol = subcol.split(root)[1].strip()
                    if ':' in rest_subcol:
                        _, option_text = rest_subcol.split(':', 1)
                        option_text = option_text.strip()
                    else:
                        option_text = rest_subcol
                    label = f"{full_question}\n{option_text}"
                    cond = df[subcol] == 1
                    col_conditions.append((label, cond))
                    col_totals[label] = cond.sum()
    
                total_label = f"{full_question}\n总计"
                total_cond = (df[subcols] == 1).any(axis=1)
                col_conditions.append((total_label, total_cond))
                col_totals[total_label] = total_cond.sum()
                continue
    
        # === 处理单选题 ===
        if q_clean in df.columns:
            values = df[q_clean].dropna().unique()
            try:
                sorted_values = sorted(values, key=lambda x: int(re.match(r'^(\d+)', str(x)).group(1)))
            except:
                sorted_values = values
            unique_question = f"{q_clean} #{instance_id}"  # 唯一标识
            for value in sorted_values:
                label = f"{unique_question}\n{value}"
                cond = df[q_clean] == value
                col_conditions.append((label, cond))
                col_totals[label] = cond.sum()
            total_label = f"{unique_question}\n总计"
            total_cond = df[q_clean].notna()
            col_conditions.append((total_label, total_cond))
            col_totals[total_label] = total_cond.sum()
        else:
            warnings.warn(f"无效问题被跳过：{q}")

    # === 行维度条件生成 ===
    row_conditions = []
    for q_type, q in valid_rows:
        if q_type == 'multi':
            root = re.match(r'^(Q\d+\.)', q).group(1)
            subcols = sorted(multi_choice_dict[root],
                           key=lambda x: extract_subcol_number(x, root))
            for subcol in subcols:
                option = subcol.split(root)[1].strip()
                cond = df[subcol] == 1
                row_conditions.append(((q, option), cond))  # 元组格式
            total_cond = (df[subcols] == 1).any(axis=1)
            row_conditions.append(((q, '总计'), total_cond))
        else:
            # 处理单选题
            values = df[q].dropna().unique()
            try:
                sorted_values = sorted(values, key=lambda x: int(re.match(r'^(\d+)', str(x)).group(1)))
            except:
                sorted_values = values
            for value in sorted_values:
                cond = df[q] == value
                row_conditions.append(((q, str(value)), cond))  # 元组形式 (问题, 值)
            # 生成总计行
            total_cond = df[q].notna()
            row_conditions.append(((q, '总计'), total_cond))

    # === 交叉统计计算 ===
    freq_results = []
    for (r_question, r_option), r_cond in row_conditions:
        row_data = {}
        for c_label, c_cond in col_conditions:
            count = (r_cond & c_cond).sum()
            row_data[c_label] = count
        freq_results.append(row_data)

    # === 创建多级索引 ===
    index = pd.MultiIndex.from_tuples(
        [(rl[0], rl[1]) for rl, _ in row_conditions],  # 提取问题和选项
        names=['问题', '选项']
    )

    freq_df = pd.DataFrame(
        freq_results,
        index=index,
        columns=[cl for cl, _ in col_conditions]
    )
    
    # === 百分比计算 ===
    percent_df = freq_df.copy()
    for col in percent_df.columns:
        total = col_totals[col]
        percent_df[col] = (freq_df[col] / total).round(3)

    # === 构建最终表格 ===
    freq_df = freq_df.add_suffix("（频数）")
    percent_df = percent_df.add_suffix("（百分比）")

    columns_order = []
    for orig_col in [cl for cl, _ in col_conditions]:
        columns_order.append(f"{orig_col}（频数）")
        columns_order.append(f"{orig_col}（百分比）")

    combined_df = pd.concat([freq_df, percent_df], axis=1)[columns_order]

    # === 新增：显著性检验计算 ===
    sig_results = []
    for r_label, r_cond in row_conditions:
        row_sig = {}
        for c_label, c_cond in col_conditions:
            # 构建列联表
            a = (r_cond & c_cond).sum()
            b = (r_cond & ~c_cond).sum()
            c = (~r_cond & c_cond).sum()
            d = (~r_cond & ~c_cond).sum()
            observed = np.array([[a, b], [c, d]])
            
            # 执行检验
            p_value = perform_significance_test(observed)
            row_sig[c_label] = p_value
        sig_results.append(row_sig)
    
    sig_df = pd.DataFrame(sig_results, 
                        index=[rl for rl, _ in row_conditions],
                        columns=[cl for cl, _ in col_conditions])

    # === 新增：生成带星号标记的显著性结果 ===
    formatted_sig_df = sig_df.copy()
    for col in formatted_sig_df.columns:
        formatted_sig_df[col] = formatted_sig_df[col].apply(
            lambda p: ''.join([s for l, s in zip(sig_levels, sig_symbols) if p <= l]) + 
                f"({p:.3f})" if not pd.isna(p) else ""
        )
    
    # === 处理已有文件 ===
    if os.path.exists(output_file):
        try:
            os.remove(output_file)
        except PermissionError:
            raise PermissionError(f"请关闭正在使用的文件：{output_file}")

    # =========================================================== Excel输出 ============================================================
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # === 提前定义样式 ===
        header_fill = PatternFill(
            start_color=header_fill_color,
            end_color=header_fill_color,
            fill_type="solid"
        )
        header_font = Font(
            name=header_font_name,
            size=header_font_size,
            bold=True,
            color=header_font_color
        )
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # === 交叉分析sheet ===
        combined_df.to_excel(writer, sheet_name='交叉分析', merge_cells=True)
        worksheet = writer.sheets['交叉分析']
        
        # === 设置百分比格式 ===
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col_idx).value
            if cell_value and "百分比" in cell_value:
                col_letter = get_column_letter(col_idx)
                for row in worksheet.iter_rows(
                    min_row=2, 
                    max_row=worksheet.max_row,
                    min_col=col_idx,
                    max_col=col_idx
                ):
                    for cell in row:
                        cell.number_format = percent_format
       
        # 标题样式
        worksheet.row_dimensions[1].height = header_height

        # 设置其他行行高（例如设置为20）
        for row_idx in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].height = 20
        for cell in worksheet[1]:
            cell.font = Font(
                name=header_font_name,
                size=header_font_size,
                bold=True,
                color=header_font_color
            )

        # === 数据条设置===
        freq_rule = DataBarRule(
            start_type='num', 
            start_value=0,
            end_type='max', 
            color=freq_data_bar_color,
            showValue="None",
            minLength=data_bar_min_length,
            maxLength=data_bar_max_length
        )
        percent_rule = DataBarRule(
            start_type='num', 
            start_value=0,
            end_type='max', 
            color=percent_data_bar_color,
            showValue="None",
            minLength=data_bar_min_length,
            maxLength=data_bar_max_length
        )

        # 预先生成有效行列表
        valid_rows = []
        for row_idx in range(2, worksheet.max_row + 1):
            row_label = worksheet.cell(row=row_idx, column=1).value or ""
            # 清理标签中的换行符和空格
            clean_label = row_label.replace('\n', '').replace(' ', '')
            # 判断是否为总计行（支持多种格式）
            if any(keyword in clean_label for keyword in ["总计", "Total", "合计"]):
                continue
            valid_rows.append(row_idx)

        # 应用数据条到所有有效列
        for col_idx in range(1, worksheet.max_column + 1):
            header_cell = worksheet.cell(row=1, column=col_idx)
            header_value = header_cell.value or ""
            
            # 确定规则类型
            if "频数" in header_value:
                rule = freq_rule
            elif "百分比" in header_value:
                rule = percent_rule
            else:
                continue

            # 仅当存在有效数据行时应用
            if valid_rows:
                col_letter = get_column_letter(col_idx)
                data_range = f"{col_letter}{min(valid_rows)}:{col_letter}{max(valid_rows)}"
                worksheet.conditional_formatting.add(data_range, rule)

        # === 格式优化 ===
        # 设置列宽
        worksheet.column_dimensions['A'].width = 25  # 问题列宽
        worksheet.column_dimensions['B'].width = 25  # 选项列宽
        
        # 设置C列及之后的宽度
        for col in worksheet.columns:
            col_letter = get_column_letter(col[0].column)
            
            # 跳过已设置的A、B列
            if col_letter in ['A', 'B']: 
                continue
                
            # 计算最大列宽
            max_length = 0
            for cell in col:
                try:
                    # 处理换行文本：取最长行的长度
                    if cell.value and '\n' in str(cell.value):
                        line_lengths = [len(line) for line in str(cell.value).split('\n')]
                        cell_length = max(line_lengths)
                    else:
                        cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
                except:
                    pass
            
            # 设置列宽（使用自定义宽度或自动调整）
            if data_column_width:  # 如果设置了固定宽度
                worksheet.column_dimensions[col_letter].width = data_column_width
            else:  # 否则自动调整宽度
                adjusted_width = min(max_length + 2, max_column_width)
                worksheet.column_dimensions[col_letter].width = adjusted_width

        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(
                    wrap_text=True, 
                    vertical='top',
                    horizontal='left'
                )

        # === 设置全局字体 ===
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = Font(name="微软雅黑")  # 保留原有其他属性

        # === 冻结前两列和第一行 ===
        worksheet.freeze_panes = "C2"
        worksheet.sheet_view.showGridLines = False   # 隐藏网格线

        # 新增显著性检验sheet
        sig_df.to_excel(writer, sheet_name='显著性检验')
        formatted_sig_df.to_excel(writer, sheet_name='带星号显著性')

        # 设置显著性sheet样式（复用已定义的样式变量）
        for sheet_name in ['显著性检验', '带星号显著性']:
            sheet = writer.sheets[sheet_name]
            for cell in sheet[1]:  # 设置标题行样式
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
            # 设置数字格式
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row[1:]:
                    if sheet_name == '显著性检验':
                        cell.number_format = '0.000'
                    cell.alignment = Alignment(horizontal='center')

    return combined_df, sig_df